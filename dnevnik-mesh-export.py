#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Выгрузка всех оценок из МЭШ (school.mos.ru) в XLSX.

Скрипт сам запускает Chrome, получает токен и cookies через CDP,
определяет список всех детей из профиля, собирает оценки каждого
за 38 учебных недель и сохраняет grades.xlsx.
"""

import io
import json
import os
import sys
import time
import traceback
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.error import URLError

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def wait_exit():
    try:
        input("Нажмите Enter для выхода...")
    except (EOFError, KeyboardInterrupt):
        print()


# === Зависимости ===
try:
    from openpyxl import Workbook
except ImportError:
    print("Ошибка: установите openpyxl -> pip install openpyxl")
    wait_exit(); sys.exit(1)

try:
    import websocket
except ImportError:
    print("Ошибка: установите websocket-client -> pip install websocket-client")
    wait_exit(); sys.exit(1)

# === Конфигурация ===
API_URL = "https://school.mos.ru"
CHROME_DEBUG_URL = "http://127.0.0.1:9222"
WEEK_COUNT = 38

_now = datetime.now()
_year = _now.year if _now.month >= 9 else _now.year - 1
FIRST_WEEK = f"{_year}-09-01"
MES_PAGE = f"https://school.mos.ru/diary/marks/current-marks?date=01.09.{_year}"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "grades.xlsx")
CHROME_PROFILE_DIR = os.path.join(SCRIPT_DIR, "chrome-profile")


# === Логгер ===
_LOG_FILE = None

def _ensure_log():
    global _LOG_FILE
    if _LOG_FILE is None:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        _LOG_FILE = os.path.join(SCRIPT_DIR, f"dnevnik-mesh-export_{ts}.log")
    return _LOG_FILE

def log_exception(context=""):
    tb = traceback.format_exc()
    path = _ensure_log()
    try:
        with open(path, "a", encoding="utf-8") as f:
            f.write(f"=== {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
            if context:
                f.write(f"Контекст: {context}\n")
            f.write(tb)
            f.write("\n\n")
    except Exception:
        pass


# === Chrome CDP ===
def find_page_ws(log_on_fail=False):
    try:
        resp = urlopen(f"{CHROME_DEBUG_URL}/json/list", timeout=5)
        pages = json.loads(resp.read().decode())
    except Exception as e:
        if log_on_fail:
            log_exception(f"find_page_ws: подключение к {CHROME_DEBUG_URL}")
        print(f"  Chrome недоступен на порту 9222: {e}")
        return None, None

    try:
        for page in pages:
            url = page.get("url", "")
            if "school.mos.ru" in url and "marks" in url and "about:blank" not in url:
                return page["webSocketDebuggerUrl"], page["id"]
        for page in pages:
            url = page.get("url", "")
            if url and url != "about:blank" and url != "chrome://welcome/":
                return page["webSocketDebuggerUrl"], page["id"]
    except Exception as e:
        if log_on_fail:
            log_exception("find_page_ws: обработка списка страниц")
        print(f"  Ошибка при разборе страниц: {e}")

    return None, None


def cdp_call(ws, method, params=None, timeout=10):
    cmd_id = int(time.time() * 1000) % 100000
    cmd = {"id": cmd_id, "method": method, "params": params or {}}
    try:
        ws.send(json.dumps(cmd))
    except Exception as e:
        log_exception(f"cdp_call({method}): ошибка отправки")
        raise

    start = time.time()
    while time.time() - start < timeout:
        try:
            raw = ws.recv()
            resp = json.loads(raw)
            if resp.get("id") == cmd_id:
                return resp
        except websocket.WebSocketTimeoutException:
            raise TimeoutError(f"CDP timeout: {method}")
        except json.JSONDecodeError:
            continue
        except Exception as e:
            log_exception(f"cdp_call({method}): ошибка получения")
            raise

    raise TimeoutError(f"CDP timeout: {method}")


def get_cookies_via_cdp(ws):
    try:
        resp = cdp_call(ws, "Network.getAllCookies", {})
        cookies = resp.get("result", {}).get("cookies", [])
    except Exception as e:
        log_exception("get_cookies_via_cdp: Network.getAllCookies")
        print(f"  Ошибка получения cookie: {e}")
        return {}

    cookies_map = {}
    try:
        for c in cookies:
            domain = c.get("domain", "")
            if "mos.ru" in domain or "school.mos.ru" in domain:
                cookies_map[c["name"]] = c["value"]
    except Exception as e:
        log_exception("get_cookies_via_cdp: обработка cookie")
    return cookies_map


# === API вызовы ===
def get_profile(token, cookie_str):
    """Получить профиль родителя и список детей."""
    url = f"{API_URL}/api/family/web/v1/profile"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}",
        "X-Mes-Subsystem": "familyweb",
        "Referer": MES_PAGE,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Cookie": cookie_str,
    }
    try:
        resp = urlopen(Request(url, headers=headers), timeout=15)
        data = json.loads(resp.read().decode())
        children = data.get("children", [])
        return children
    except Exception as e:
        log_exception("get_profile")
        print(f"  Ошибка получения профиля: {e}")
        return []


def get_marks(token, cookie_str, student_id, from_date, to_date):
    url = f"{API_URL}/api/family/web/v1/marks?student_id={student_id}&from={from_date}&to={to_date}"
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "X-Mes-Subsystem": "familyweb",
        "Referer": MES_PAGE,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Cookie": cookie_str,
    }
    req = Request(url, headers=headers, method="GET")
    try:
        resp = urlopen(req, timeout=15)
        data = json.loads(resp.read().decode())
        return data.get("payload", [])
    except URLError as e:
        log_exception(f"get_marks({from_date}..{to_date}): HTTP error")
        print(f"  HTTP ошибка {from_date}..{to_date}: {e}")
        return []
    except json.JSONDecodeError as e:
        log_exception(f"get_marks({from_date}..{to_date}): JSON decode")
        print(f"  JSON ошибка {from_date}..{to_date}: {e}")
        return []
    except Exception as e:
        log_exception(f"get_marks({from_date}..{to_date})")
        return []


def fmt_date(d):
    return d.strftime("%d.%m.%Y")


# === Автозапуск Chrome ===
def find_chrome_exe():
    import subprocess
    try:
        result = subprocess.run(["where", "chrome.exe"], capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            path = result.stdout.strip().split("\n")[0].strip()
            if os.path.exists(path):
                return path
    except Exception:
        pass
    candidates = [
        os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def start_chrome_with_debug():
    import subprocess, time
    print("  Пытаюсь автоматически запустить Chrome...")
    chrome_path = find_chrome_exe()
    if not chrome_path:
        print("  Chrome не найден на этом компьютере.")
        return False
    print(f"  Chrome найден: {chrome_path}")
    print("  Закрываю старый Chrome...")
    try:
        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], capture_output=True, timeout=10)
        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], capture_output=True, timeout=10)
        time.sleep(2)
    except Exception:
        pass
    print("  Запускаю Chrome с debug-портом...")
    try:
        subprocess.Popen(
            [chrome_path, "--remote-debugging-port=9222",
             "--remote-allow-origins=*",
             "--no-first-run",
             "--no-default-browser-check",
             "--disable-features=ChromeWhatsNewUI",
             f"--user-data-dir={CHROME_PROFILE_DIR}",
             MES_PAGE],
            shell=False,
        )
        time.sleep(1)
    except Exception as e:
        log_exception("start_chrome_with_debug: Popen")
        print(f"  Ошибка запуска: {e}")
        return False
    print("  Ожидание порта 9222...", end="", flush=True)
    for i in range(15):
        time.sleep(1)
        try:
            resp = urlopen("http://127.0.0.1:9222/json/version", timeout=2)
            print(f" готово ({i+1} сек)")
            return True
        except Exception:
            print(".", end="", flush=True)
    print(" порт не открылся.")
    return False


def build_cookie_str(cookies_map):
    important = ['session-cookie', 'JSESSIONID', 'student_person_id',
                 'active_student', 'cluster_id', 'aupd_token',
                 'oxxfgh', 'uwyii', 'uwyiert']
    parts = [f"{k}={cookies_map[k]}" for k in important if k in cookies_map]
    return "; ".join(parts)


# === Главная ===
def main():
    print("=" * 60)
    print("    Выгрузка оценок из МЭШ")
    print("=" * 60)

    # ===== 1. Подключение к Chrome =====
    print("\n[1] Поиск Chrome с debug-портом 9222...")
    page_ws = None
    try:
        page_ws, page_id = find_page_ws()
    except Exception:
        pass

    if not page_ws:
        print("  Chrome не найден на порту 9222.")
        print("  Пробую запустить автоматически...")
        if not start_chrome_with_debug():
            log_exception("main: Chrome не удалось запустить")
            print("  Chrome не найден.")
            print("  1. Закройте Chrome")
            print('  2. Запустите: chrome.exe --remote-debugging-port=9222')
            print(f"  3. Зайдите на {MES_PAGE} и авторизуйтесь")
            wait_exit(); sys.exit(1)
    else:
        if page_id != "browser":
            try:
                ws_tmp = websocket.create_connection(page_ws, timeout=5)
                resp = cdp_call(ws_tmp, "Network.getAllCookies", {})
                ws_tmp.close()
                cookies = resp.get("result", {}).get("cookies", [])
                cnames = {c.get("name"): c.get("value") for c in cookies}
                if cnames.get("aupd_token") and cnames.get("active_student"):
                    print("  Chrome найден, токен есть.")
                else:
                    print("  Chrome найден, но нет полной авторизации. Жду...")
                    page_ws = None
            except Exception:
                page_ws = None

    # Ожидание авторизации
    if not page_ws:
        print("  Ожидание авторизации в МЭШ...", end="", flush=True)
        port_lost_count = 0
        for _ in range(150):
            time.sleep(2)
            print(".", end="", flush=True)
            try:
                pages_raw = urlopen("http://127.0.0.1:9222/json/list", timeout=3).read()
                pages = json.loads(pages_raw)
                port_lost_count = 0
                for p in pages:
                    url = p.get("url", "")
                    if "school.mos.ru" in url and "marks" in url and "about:blank" not in url:
                        try:
                            ws_tmp = websocket.create_connection(p["webSocketDebuggerUrl"], timeout=5)
                            resp = cdp_call(ws_tmp, "Network.getAllCookies", {})
                            ws_tmp.close()
                            cookies = resp.get("result", {}).get("cookies", [])
                            has_token = has_student = False
                            for c in cookies:
                                if c.get("name") == "aupd_token" and c.get("value"): has_token = True
                                if c.get("name") == "active_student" and c.get("value"): has_student = True
                            if has_token and has_student:
                                page_ws = p["webSocketDebuggerUrl"]
                                page_id = p["id"]
                                break
                        except Exception:
                            pass
                        break
                if page_ws:
                    print(" авторизация подтверждена!")
                    try:
                        import ctypes
                        k32 = ctypes.windll.kernel32
                        u32 = ctypes.windll.user32
                        hwnd = k32.GetConsoleWindow()
                        if hwnd:
                            u32.ShowWindow(hwnd, 9)
                            u32.SetForegroundWindow(hwnd)
                            u32.BringWindowToTop(hwnd)
                    except Exception:
                        pass
                    break
            except Exception:
                port_lost_count += 1
                # Если порт 9222 пропал (Chrome перезапустился), восстанавливаем
                if port_lost_count >= 5:
                    port_lost_count = 0
                    print("\n  Потерян debug-порт. Перезапускаю Chrome...")
                    try:
                        import subprocess
                        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"],
                                       capture_output=True, timeout=10)
                        time.sleep(1)
                        chrome_path2 = find_chrome_exe()
                        if chrome_path2:
                            subprocess.Popen(
                                [chrome_path2, "--remote-debugging-port=9222",
                                 "--remote-allow-origins=*",
                                 "--no-first-run",
                                 "--no-default-browser-check",
                                 "--disable-features=ChromeWhatsNewUI",
                                 f"--user-data-dir={CHROME_PROFILE_DIR}",
                                 MES_PAGE], shell=False)
                            time.sleep(1)
                    except Exception:
                        pass
                    print("  Ожидание авторизации в МЭШ...", end="", flush=True)

        if not page_ws:
            print("\n  Превышено время ожидания (5 мин). Авторизуйтесь в Chrome.")
            wait_exit(); sys.exit(1)

    print("  Подключение к CDP...")

    # ===== 2. WebSocket + Cookies =====
    ws = None
    try:
        ws = websocket.create_connection(page_ws, timeout=10)
    except Exception as e:
        log_exception("main[2]: WebSocket connection")
        print(f"  Ошибка подключения: {e}")
        wait_exit(); sys.exit(1)

    cookies_map = {}
    try:
        cookies_map = get_cookies_via_cdp(ws)
        print(f"  Получено cookies: {len(cookies_map)}")
    except Exception as e:
        print(f"  Ошибка получения cookies: {e}")

    token = cookies_map.get("aupd_token", "")
    if not token:
        print("  Ошибка: aupd_token не найден.")
        ws.close(); wait_exit(); sys.exit(1)

    try:
        ws.close()
    except Exception:
        pass

    cookie_str = build_cookie_str(cookies_map)
    print(f"  Token: {token[:30]}...{token[-10:]}")

    # ===== 3. Получение списка детей =====
    print("\n[2] Получение списка детей из профиля...")
    children = get_profile(token, cookie_str)

    if not children:
        print("  Не удалось получить список детей.")
        wait_exit(); sys.exit(1)

    print(f"  Найдено детей: {len(children)}")
    for ch in children:
        print(f"    - {ch.get('last_name','')} {ch.get('first_name','')} (ID: {ch.get('id')}, {ch.get('class_name','')} класс)")

    # ===== 4. Сбор оценок =====
    print(f"\n[3] Сбор оценок за {WEEK_COUNT} недель...")
    print("-" * 60)

    start = datetime.strptime(FIRST_WEEK, "%Y-%m-%d")
    all_marks = []
    weeks_with_data = 0
    weeks_empty = 0
    total_weeks = 0

    for child in children:
        student_id = child["id"]
        child_name = f"{child.get('last_name', '')} {child.get('first_name', '')}".strip()
        print(f"\n  Ребёнок: {child_name} (ID: {student_id})")

        for wn in range(1, WEEK_COUNT + 1):
            try:
                ws_ = start + timedelta(weeks=wn - 1)
                we_ = ws_ + timedelta(days=6)
                f_str = ws_.strftime("%Y-%m-%d")
                t_str = we_.strftime("%Y-%m-%d")
                total_weeks += 1

                marks = get_marks(token, cookie_str, student_id, f_str, t_str)

                if marks:
                    weeks_with_data += 1
                    print(f"    Неделя {wn:2d} ({f_str}..{t_str}): {len(marks)} оценок")
                    for m in marks:
                        try:
                            date_raw = m.get("date", "")
                            if date_raw:
                                d = datetime.strptime(date_raw, "%Y-%m-%d")
                                all_marks.append({
                                    "child": child_name,
                                    "date": fmt_date(d),
                                    "subject": m.get("subject_name", ""),
                                    "topic": m.get("control_form_name", ""),
                                    "grade": m.get("value", ""),
                                })
                        except Exception as e:
                            log_exception(f"main: parse mark {wn}")
                            continue
                else:
                    weeks_empty += 1
            except Exception as e:
                log_exception(f"main: week {wn}")
                print(f"    Неделя {wn:2d}: Ошибка: {e}")
                weeks_empty += 1

    print("-" * 60)

    if not all_marks:
        print("\nНет данных. Проверьте авторизацию в МЭШ.")
        wait_exit(); sys.exit(1)

    # ===== 5. Сохранение =====
    print(f"\n[4] Сохранение в {OUTPUT_FILE}...")

    if os.path.exists(OUTPUT_FILE):
        try:
            os.remove(OUTPUT_FILE)
        except Exception:
            pass

    try:
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        from collections import defaultdict

        wb = Workbook()
        ws = wb.active
        ws.title = "Оценки"

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # === Основная таблица ===
        headers_main = ["Ребёнок", "Дата", "Предмет", "Тема урока", "Оценка"]
        ws.append(headers_main)
        for c in range(1, len(headers_main) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True)
            cell.border = thin_border

        for m in all_marks:
            r = ws.max_row + 1
            for c, val in enumerate([m["child"], m["date"], m["subject"], m["topic"], m["grade"]], 1):
                cell = ws.cell(r, c, val)
                cell.border = thin_border

        # Автоширина основной таблицы
        for ci in range(1, len(headers_main) + 1):
            ml = len(headers_main[ci - 1])
            for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
                for cell in row:
                    if cell:
                        try:
                            ml = max(ml, len(str(cell)))
                        except Exception:
                            pass
            ws.column_dimensions[chr(64 + ci)].width = min(ml + 3, 60)

        # === Сводные таблицы (по каждому ребёнку, одна под другой) ===
        from collections import defaultdict

        # Группируем оценки по детям
        marks_by_child = defaultdict(list)
        for m in all_marks:
            marks_by_child[m["child"]].append(m)

        sc = len(headers_main) + 3  # стартовая колонка сводных
        current_row = 1

        for child_name in sorted(marks_by_child.keys()):
            child_marks = marks_by_child[child_name]

            # Заголовок с именем ребёнка
            cell = ws.cell(current_row, sc, child_name)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="left")
            current_row += 1

            # Собираем pivot для этого ребёнка
            pivot = defaultdict(lambda: defaultdict(int))
            subjects = set()
            grades = set()
            for m in child_marks:
                pivot[m["subject"]][m["grade"]] += 1
                subjects.add(m["subject"])
                grades.add(m["grade"])

            subjects = sorted(subjects)
            grade_order = sorted(grades, key=lambda x: (x.isdigit() == False, int(x) if x.isdigit() else x))

            # Заголовок: Предмет | оценка1 | ... | Итого
            headers_pivot = ["Предмет"] + grade_order + ["Итого"]
            for ci, h in enumerate(headers_pivot, sc):
                cell = ws.cell(current_row, ci, h)
                cell.font = Font(bold=True)
                cell.border = thin_border
            current_row += 1

            # Данные
            for subj in subjects:
                ws.cell(current_row, sc, subj).border = thin_border
                row_total = 0
                for gi, grade in enumerate(grade_order, sc + 1):
                    val = pivot[subj].get(grade, 0)
                    cell = ws.cell(current_row, gi, val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    row_total += val
                cell_total = ws.cell(current_row, sc + len(grade_order) + 1, row_total)
                cell_total.border = thin_border
                cell_total.font = Font(bold=True)
                cell_total.alignment = Alignment(horizontal="center")
                current_row += 1

            # Итого строка
            cell = ws.cell(current_row, sc, "Итого")
            cell.font = Font(bold=True)
            cell.border = thin_border
            grand_total = 0
            for gi, grade in enumerate(grade_order, sc + 1):
                col_total = sum(pivot[s].get(grade, 0) for s in subjects)
                cell = ws.cell(current_row, gi, col_total)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                grand_total += col_total
            cell_gt = ws.cell(current_row, sc + len(grade_order) + 1, grand_total)
            cell_gt.font = Font(bold=True)
            cell_gt.border = thin_border
            cell_gt.alignment = Alignment(horizontal="center")

            current_row += 2  # пустая строка между детьми

        # Автоширина колонок сводных
        ws.column_dimensions[chr(64 + sc)].width = max(35, len("Предмет") + 3)
        # Остальные колонки (оценки + Итого) выставляем по максимальному children
        max_grades = max(len(marks_by_child[c]) for c in marks_by_child)  # не то
        max_grade_count = 0
        for cm in marks_by_child.values():
            gs = set(m["grade"] for m in cm)
            max_grade_count = max(max_grade_count, len(gs))
        max_grade_count += 1  # +1 для Итого
        for gi in range(max_grade_count + 1):
            col_idx = sc + 1 + gi
            if col_idx <= 90:
                ws.column_dimensions[chr(64 + col_idx)].width = 10

        # === Автофильтр на основную таблицу ===
        last_main_row = len(all_marks) + 1
        ws.auto_filter.ref = f"A1:E{last_main_row}"

        wb.save(OUTPUT_FILE)
        save_ok = True
    except Exception as e:
        log_exception("main[6]: save XLSX")
        print(f"\nОшибка сохранения файла: {e}")
        save_ok = False

    # ===== Закрытие Chrome =====
    print("\nЗакрываю Chrome...")
    try:
        import subprocess
        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], capture_output=True, timeout=10)
        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], capture_output=True, timeout=10)
        print("  Chrome закрыт.")
    except Exception:
        pass

    # ===== Статистика =====
    print(f"\n{'=' * 60}")
    print("           С Т А Т И С Т И К А")
    print(f"{'=' * 60}")
    print(f"  Детей:                     {len(children)}")
    print(f"  Всего оценок собрано:      {len(all_marks)}")
    print(f"  Недель с данными:          {weeks_with_data}")
    print(f"  Всего недель проверено:    {total_weeks}")
    if weeks_with_data:
        print(f"  Среднее оценок за неделю:  {len(all_marks)//max(weeks_with_data,1):.1f}")
    print()
    if save_ok:
        print(f"  Файл сохранён: {OUTPUT_FILE}")
    else:
        print(f"  Файл НЕ сохранён из-за ошибки!")
    if _LOG_FILE and os.path.exists(_LOG_FILE):
        print(f"  Лог-файл:     {_LOG_FILE}")
    else:
        print("  Лог: ошибок не было")
    print(f"{'=' * 60}")
    print()
    wait_exit()


if __name__ == "__main__":
    main()
